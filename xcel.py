# This program makes use of the openpyxl module
# to open an excel workbook and use values within it

import openpyxl

# open the workbook using the load_workbook function and save it to a variable
workbook = openpyxl.load_workbook('C:\\Users\\JJaimeRivera\\pythonScripts\\automate_tutorials\\example.xlsx')
#print(type(workbook))

# get the sheet tiled 'Sheet1' from the workbook
sheet = workbook['Sheet1']

#if you do not know the sheet names, this line can retrieve them
workbook.sheetnames

#get the cell A1 from the sheet
cell = sheet['A1']

#get the cell value
cell.value

#get the cell value as a string (previously in date/time format)
str(cell.value)

#iterate through the cells in the second column and print each value
for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
    
